#!/usr/bin/env python3
"""Headless-callable CLI for work-hours-log-automation.

This is the ONE surface both the interactive Claude Code skill and any future
scheduler are meant to call — plain argparse, no interactive prompts, JSON on
stdout. See references/cli-reference.md for the full command/argument/exit-code
reference and the deferred-scheduling call shape.

Exit codes: 0 success, 1 validation/input error, 2 locked period without a valid
override, 3 file/path not found, 4 workbook was changed by someone else since it
was loaded (concurrent-modification guard, see workbook_builder.py).
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import date, datetime, time
from pathlib import Path
from typing import List, Optional

from models import AttendanceRecord, DataSource, EmployeeStatus
from employee_master import EmployeeMaster
from holiday_calendar import HolidayCalendar
from csv_importer import CsvImporter
from business_rules import BusinessRulesEngine
from workbook_builder import WorkbookBuilder, ConcurrentModificationError
from period_lock import PeriodLockManager, LockedPeriodError
from location_config import LocationConfig
import audit_log


def _emit(payload: dict, code: int = 0) -> int:
    print(json.dumps(payload, indent=2, default=str))
    return code


def _parse_time_or_none(value) -> Optional[time]:
    if value in (None, ""):
        return None
    return datetime.strptime(value, "%H:%M").time()


def _load_entries_json(path: Path) -> List[AttendanceRecord]:
    rows = json.loads(path.read_text(encoding="utf-8"))
    records = []
    for row in rows:
        status = EmployeeStatus.LEAVE if row.get("employeeStatus") == "Leave" else EmployeeStatus.PRESENT
        records.append(
            AttendanceRecord(
                work_date=datetime.strptime(row["workDate"], "%Y-%m-%d").date(),
                employee_id=row["employeeId"],
                arrival_time=_parse_time_or_none(row.get("arrivalTime")),
                departure_time=_parse_time_or_none(row.get("departureTime")),
                employee_status=status,
                data_source=DataSource.MANUAL,
                notes=row.get("notes", ""),
            )
        )
    return records


def _workbook_path(workbook_dir: Path, month: str) -> Path:
    return workbook_dir / f"{month}-work-hours-log.xlsx"


def _validate_month(month: str) -> None:
    datetime.strptime(month, "%Y-%m")


# ---- generate ---------------------------------------------------------------


def cmd_generate(args: argparse.Namespace) -> int:
    try:
        _validate_month(args.month)
    except ValueError:
        return _emit({"error": f"--month {args.month!r} is not YYYY-MM"}, 1)

    try:
        employee_master = EmployeeMaster.load(args.employee_master)
        holiday_calendar = HolidayCalendar.load(args.holiday_config)
    except FileNotFoundError as exc:
        return _emit({"error": str(exc)}, 3)
    except ValueError as exc:
        return _emit({"error": str(exc)}, 1)

    if not args.entries_json and not args.csv:
        return _emit({"error": "Provide --entries-json and/or --csv as input"}, 1)

    records: List[AttendanceRecord] = []
    warnings: List[str] = []

    if args.entries_json:
        try:
            records.extend(_load_entries_json(args.entries_json))
        except FileNotFoundError:
            return _emit({"error": f"--entries-json not found: {args.entries_json}"}, 3)

    if args.csv:
        if not args.mapping_json:
            return _emit({"error": "--mapping-json is required when --csv is given"}, 1)
        try:
            importer = CsvImporter.load_mapping(args.mapping_json)
        except FileNotFoundError:
            return _emit({"error": f"--mapping-json not found: {args.mapping_json}"}, 3)
        except ValueError as exc:
            return _emit({"error": str(exc)}, 1)
        if not args.csv.exists():
            return _emit({"error": f"--csv not found: {args.csv}"}, 3)
        csv_records, csv_warnings = importer.import_rows(args.csv, employee_master)
        records.extend(csv_records)
        warnings.extend(csv_warnings)

    in_month, out_of_month = [], []
    for record in records:
        (in_month if record.work_date.strftime("%Y-%m") == args.month else out_of_month).append(record)
    for record in out_of_month:
        warnings.append(
            f"employee {record.employee_id!r} date {record.work_date} is outside "
            f"target month {args.month!r}, skipped"
        )

    lock_mgr = PeriodLockManager(args.lock_state)
    try:
        override_valid = lock_mgr.require_override_if_locked(
            args.month, args.override_name, args.override_reason
        )
    except LockedPeriodError as exc:
        return _emit({"error": str(exc), "month": args.month, "locked": True}, 2)

    workbook_path = _workbook_path(args.workbook_dir, args.month)
    builder = WorkbookBuilder(workbook_path)
    builder.load_or_create()

    prefilled_count = builder.prefill_month(args.month, employee_master, holiday_calendar)
    upsert_result = builder.upsert_attendance_rows(in_month, employee_master, holiday_calendar)
    warnings.extend(upsert_result.warnings)

    engine = BusinessRulesEngine(holiday_calendar)
    calculated_rows = builder.rebuild_calculation_sheet(engine, employee_master)
    summary_tables = engine.summarize(calculated_rows, employee_master)
    builder.rebuild_summary_sheet(summary_tables)
    builder.write_front_matter(args.month, employee_master)
    builder.apply_data_validation_and_formatting()

    needs_review = [
        {"employeeId": r.employee_id, "workDate": r.work_date.isoformat()}
        for r in calculated_rows
        if r.review_status.value == "Needs Supervisor Review"
    ]

    result = {
        "month": args.month,
        "workbookPath": str(workbook_path),
        "dryRun": args.dry_run,
        "prefilledBlankRows": prefilled_count,
        "added": upsert_result.added,
        "updated": upsert_result.updated,
        "warnings": warnings,
        "needsSupervisorReview": needs_review,
        "lockOverrideApplied": bool(override_valid),
    }

    if args.dry_run:
        result["saved"] = False
        result["backupPath"] = None
        return _emit(result, 0)

    try:
        builder.raise_if_modified_since_load()
    except ConcurrentModificationError as exc:
        return _emit({"error": str(exc), "month": args.month, "concurrentModification": True}, 4)

    backup_path = builder.backup_if_exists(args.workbook_dir / "backups")
    builder.save()

    audit_fields = {
        "Operator": args.operator_name,
        "Month": args.month,
        "Action": "generate",
        "Added": upsert_result.added,
        "Updated": upsert_result.updated,
    }
    if override_valid:
        audit_fields["Locked override"] = "Yes"
        audit_fields["Override reason"] = args.override_reason
    audit_log.append_entry(args.audit_log, audit_fields)

    result["saved"] = True
    result["backupPath"] = str(backup_path) if backup_path else None
    return _emit(result, 0)


# ---- import-csv-preview ------------------------------------------------------


def cmd_import_csv_preview(args: argparse.Namespace) -> int:
    try:
        employee_master = EmployeeMaster.load(args.employee_master)
    except FileNotFoundError as exc:
        return _emit({"error": str(exc)}, 3)
    except ValueError as exc:
        return _emit({"error": str(exc)}, 1)

    try:
        importer = CsvImporter.load_mapping(args.mapping_json)
    except FileNotFoundError:
        return _emit({"error": f"--mapping-json not found: {args.mapping_json}"}, 3)
    except ValueError as exc:
        return _emit({"error": str(exc)}, 1)

    if not args.csv.exists():
        return _emit({"error": f"--csv not found: {args.csv}"}, 3)

    records, warnings = importer.import_rows(args.csv, employee_master)

    normalized = [
        {
            "employeeId": r.employee_id,
            "workDate": r.work_date.isoformat(),
            "arrivalTime": r.arrival_time.strftime("%H:%M") if r.arrival_time else None,
            "departureTime": r.departure_time.strftime("%H:%M") if r.departure_time else None,
            "employeeStatus": r.employee_status.value,
            "notes": r.notes,
        }
        for r in records
    ]
    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text(json.dumps(normalized, indent=2), encoding="utf-8")

    return _emit(
        {
            "importedCount": len(records),
            "skippedCount": len(warnings),
            "warnings": warnings,
            "normalizedEntriesPath": str(args.out),
        },
        0,
    )


# ---- lock ---------------------------------------------------------------------


def cmd_lock(args: argparse.Namespace) -> int:
    try:
        _validate_month(args.month)
    except ValueError:
        return _emit({"error": f"--month {args.month!r} is not YYYY-MM"}, 1)
    lock_mgr = PeriodLockManager(args.lock_state)
    lock_mgr.lock(args.month, args.locked_by)
    return _emit({"month": args.month, "locked": True, "lockedBy": args.locked_by}, 0)


# ---- status ---------------------------------------------------------------------


def cmd_status(args: argparse.Namespace) -> int:
    try:
        _validate_month(args.month)
    except ValueError:
        return _emit({"error": f"--month {args.month!r} is not YYYY-MM"}, 1)

    lock_mgr = PeriodLockManager(args.lock_state)
    locked = lock_mgr.is_locked(args.month)

    workbook_path = _workbook_path(args.workbook_dir, args.month)
    exists = workbook_path.exists()

    review_count = None
    if exists:
        from openpyxl import load_workbook

        wb = load_workbook(str(workbook_path), read_only=True, data_only=True)
        ws = wb["Work Hours Calculation"]
        review_count = 0
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        review_col = header.index("Review Status") if "Review Status" in header else None
        if review_col is not None:
            for row in ws.iter_rows(min_row=2):
                if row[review_col].value == "Needs Supervisor Review":
                    review_count += 1

    return _emit(
        {
            "month": args.month,
            "locked": locked,
            "workbookExists": exists,
            "workbookPath": str(workbook_path),
            "needsSupervisorReviewCount": review_count,
        },
        0,
    )


# ---- set-location / get-location -----------------------------------------------


def cmd_set_location(args: argparse.Namespace) -> int:
    base_folder = args.base_folder.resolve()
    for sub in ("_config", "workbooks", "workbooks/backups"):
        (base_folder / sub).mkdir(parents=True, exist_ok=True)
    LocationConfig().save(base_folder)
    paths = LocationConfig.derive_paths(base_folder)
    return _emit({"baseFolder": str(base_folder), "paths": {k: str(v) for k, v in paths.items()}}, 0)


def cmd_get_location(args: argparse.Namespace) -> int:
    base_folder = LocationConfig().load()
    if base_folder is None:
        return _emit({"baseFolder": None, "paths": None}, 0)
    paths = LocationConfig.derive_paths(base_folder)
    return _emit({"baseFolder": str(base_folder), "paths": {k: str(v) for k, v in paths.items()}}, 0)


# ---- argparse wiring ------------------------------------------------------------


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(prog="work-hours-log-automation")
    sub = parser.add_subparsers(dest="command", required=True)

    g = sub.add_parser("generate")
    g.add_argument("--employee-master", required=True, type=Path)
    g.add_argument("--holiday-config", required=True, type=Path)
    g.add_argument("--month", required=True)
    g.add_argument("--workbook-dir", required=True, type=Path)
    g.add_argument("--lock-state", required=True, type=Path)
    g.add_argument("--audit-log", required=True, type=Path)
    g.add_argument("--operator-name", required=True)
    g.add_argument("--entries-json", type=Path)
    g.add_argument("--csv", type=Path)
    g.add_argument("--mapping-json", type=Path)
    g.add_argument("--override-name")
    g.add_argument("--override-reason")
    g.add_argument("--dry-run", action="store_true")
    g.set_defaults(func=cmd_generate)

    p = sub.add_parser("import-csv-preview")
    p.add_argument("--csv", required=True, type=Path)
    p.add_argument("--mapping-json", required=True, type=Path)
    p.add_argument("--employee-master", required=True, type=Path)
    p.add_argument("--out", required=True, type=Path)
    p.set_defaults(func=cmd_import_csv_preview)

    l = sub.add_parser("lock")
    l.add_argument("--month", required=True)
    l.add_argument("--locked-by", required=True)
    l.add_argument("--lock-state", required=True, type=Path)
    l.set_defaults(func=cmd_lock)

    s = sub.add_parser("status")
    s.add_argument("--month", required=True)
    s.add_argument("--workbook-dir", required=True, type=Path)
    s.add_argument("--lock-state", required=True, type=Path)
    s.set_defaults(func=cmd_status)

    sl = sub.add_parser("set-location")
    sl.add_argument("--base-folder", required=True, type=Path)
    sl.set_defaults(func=cmd_set_location)

    gl = sub.add_parser("get-location")
    gl.set_defaults(func=cmd_get_location)

    return parser


def main(argv=None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
