"""Builds/updates the 3-sheet monthly work-hours workbook with openpyxl.

Sheet 1 (Attendance Log) is the only sheet ever hand-edited (via upsert — existing
Arrival/Departure values are never blanked by an incoming record that doesn't
supply them). Sheets 2 and 3 are fully derived: every call clears and rewrites
them from Sheet 1's current contents, so they can never drift out of sync.

Sheet 1 and Sheet 2 carry a title banner and a short front-matter block (Sheet 1
also shows departments/month/break as read-only, generated FROM the config —
never the source of truth; the JSON config stays authoritative). Sheet 3 keeps
its existing plain layout unchanged, by explicit choice.
"""

from __future__ import annotations

import calendar
import hashlib
import shutil
from dataclasses import dataclass, field
from datetime import date, datetime, time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.formatting.rule import CellIsRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from models import (
    AttendanceRecord,
    CalculatedRow,
    DataSource,
    DayType,
    EmployeeStatus,
    ReviewStatus,
    SummaryTables,
)
from employee_master import EmployeeMaster
from holiday_calendar import HolidayCalendar
from business_rules import BusinessRulesEngine, BREAK_HOURS


class ConcurrentModificationError(Exception):
    """Raised when the workbook on disk changed after it was loaded into memory
    (e.g. another user on another PC saved it first) — the caller must stop
    without writing, never silently overwrite the other change."""


SHEET_ATTENDANCE = "Attendance Log"
SHEET_CALCULATION = "Work Hours Calculation"
SHEET_SUMMARY = "Summary Report"

ATTENDANCE_HEADERS = (
    "Work Date", "Day Type", "Employee ID", "Employee Name", "Department",
    "Shift Start", "Shift End", "Arrival Time", "Departure Time",
    "Employee Status", "Data Source", "Review Status", "Notes",
)

CALCULATION_HEADERS = (
    "Work Date", "Employee ID", "Employee Name", "Day Type", "Employee Status",
    "Arrival Time", "Departure Time", "Gross Hours", "Break Hours",
    "Confirmed Work Hours", "Expected Hours", "Late Arrival (Hours)",
    "Early Departure (Hours)", "Review Status",
)

DAILY_HEADERS = ("Employee ID", "Employee Name", "Work Date", "Daily Work Hours")
WEEKLY_HEADERS = ("Employee ID", "Employee Name", "Week Start", "Week End", "Weekly Work Hours")
MONTHLY_HEADERS = ("Employee ID", "Employee Name", "Month", "Monthly Work Hours")

# Front-matter layout (banner + metadata block) for Sheet 1 and Sheet 2.
BANNER_FILL = PatternFill(start_color="FF17365D", end_color="FF17365D", fill_type="solid")
BANNER_FONT = Font(bold=True, color="FFFFFFFF", size=12)
METADATA_FILL = PatternFill(start_color="FFEAF3F8", end_color="FFEAF3F8", fill_type="solid")
REVIEW_HIGHLIGHT_FILL = PatternFill(start_color="FFFFC7CE", end_color="FFFFC7CE", fill_type="solid")
REVIEW_HIGHLIGHT_FONT = Font(color="FF9C0006")

ATTENDANCE_TITLE_ROW = 1
ATTENDANCE_DEPARTMENTS_ROW = 3
ATTENDANCE_MONTH_ROW = 4
ATTENDANCE_BREAK_ROW = 5
ATTENDANCE_NOTE_ROW = 7
ATTENDANCE_HEADER_ROW = 9
ATTENDANCE_FIRST_DATA_ROW = 10

CALCULATION_TITLE_ROW = 1
CALCULATION_NOTE_ROW = 3
CALCULATION_HEADER_ROW = 5
CALCULATION_FIRST_DATA_ROW = 6


@dataclass
class UpsertResult:
    added: int = 0
    updated: int = 0
    preserved_unchanged: int = 0
    touched: List[Tuple[str, str]] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


def _write_header(ws: Worksheet, headers: Tuple[str, ...], row: int) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True)


def _col_index(headers: Tuple[str, ...], name: str) -> int:
    return headers.index(name) + 1  # 1-based for openpyxl


def _time_or_none(value) -> Optional[time]:
    if value is None or value == "":
        return None
    if isinstance(value, time):
        return value
    if isinstance(value, datetime):
        return value.time()
    return None


def _date_or_none(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    return None


def _write_banner(ws: Worksheet, title: str, num_cols: int, row: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = BANNER_FONT
    cell.fill = BANNER_FILL
    for c in range(1, num_cols + 1):
        ws.cell(row=row, column=c).fill = BANNER_FILL


def _write_note(ws: Worksheet, text: str, num_cols: int, row: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.fill = METADATA_FILL
    for c in range(1, num_cols + 1):
        ws.cell(row=row, column=c).fill = METADATA_FILL


def _write_metadata_row(ws: Worksheet, label: str, value, row: int) -> None:
    label_cell = ws.cell(row=row, column=1, value=label)
    label_cell.font = Font(bold=True)
    label_cell.fill = METADATA_FILL
    value_cell = ws.cell(row=row, column=2, value=value)
    value_cell.fill = METADATA_FILL


class WorkbookBuilder:
    def __init__(self, workbook_path: Path):
        self.workbook_path = Path(workbook_path)
        self._wb: Optional[Workbook] = None
        self._loaded_hash: Optional[str] = None  # None means "did not exist at load time"

    @staticmethod
    def _file_hash(path: Path) -> str:
        return hashlib.sha256(path.read_bytes()).hexdigest()

    def load_or_create(self) -> None:
        if self.workbook_path.exists():
            self._loaded_hash = self._file_hash(self.workbook_path)
            self._wb = load_workbook(str(self.workbook_path))
            for name in (SHEET_ATTENDANCE, SHEET_CALCULATION, SHEET_SUMMARY):
                if name not in self._wb.sheetnames:
                    self._wb.create_sheet(name)
            self._ensure_attendance_layout()
            self._ensure_calculation_layout()
        else:
            self._loaded_hash = None
            self._wb = Workbook()
            default_sheet = self._wb.active
            self._wb.remove(default_sheet)
            self._wb.create_sheet(SHEET_ATTENDANCE)
            self._wb.create_sheet(SHEET_CALCULATION)
            self._wb.create_sheet(SHEET_SUMMARY)
            self._ensure_attendance_layout()
            self._ensure_calculation_layout()

    def _ensure_attendance_layout(self) -> None:
        ws = self._wb[SHEET_ATTENDANCE]
        num_cols = len(ATTENDANCE_HEADERS)
        if ws.cell(row=ATTENDANCE_HEADER_ROW, column=1).value != ATTENDANCE_HEADERS[0]:
            _write_header(ws, ATTENDANCE_HEADERS, ATTENDANCE_HEADER_ROW)
            for col_letter, width in zip("ABCDEFGHIJKLM", (13, 15, 13, 18, 14, 11, 11, 12, 13, 14, 12, 22, 24)):
                ws.column_dimensions[col_letter].width = width
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = f"A{ATTENDANCE_HEADER_ROW + 1}"

    def _ensure_calculation_layout(self) -> None:
        ws = self._wb[SHEET_CALCULATION]
        num_cols = len(CALCULATION_HEADERS)
        if ws.cell(row=CALCULATION_HEADER_ROW, column=1).value != CALCULATION_HEADERS[0]:
            _write_header(ws, CALCULATION_HEADERS, CALCULATION_HEADER_ROW)
            for col_letter, width in zip(
                "ABCDEFGHIJKLMN", (13, 13, 15, 13, 15, 11, 11, 11, 11, 18, 14, 16, 18, 22)
            ):
                ws.column_dimensions[col_letter].width = width
            ws.sheet_view.showGridLines = False
            ws.freeze_panes = f"A{CALCULATION_HEADER_ROW + 1}"

    def write_front_matter(
        self, month: str, employee_master: EmployeeMaster
    ) -> None:
        """Read-only, generated-from-config display — never the source of truth.
        Safe to call every run; always overwrites with the current config."""
        departments = sorted({e.department for e in employee_master.all_active()})
        dept_display = ", ".join(departments) if departments else "(none configured)"

        ws1 = self._wb[SHEET_ATTENDANCE]
        _write_banner(ws1, f"Work Hours Log — Attendance Log — {month}", len(ATTENDANCE_HEADERS), ATTENDANCE_TITLE_ROW)
        _write_metadata_row(ws1, "Departments in scope", dept_display, ATTENDANCE_DEPARTMENTS_ROW)
        _write_metadata_row(ws1, "Reporting Month", month, ATTENDANCE_MONTH_ROW)
        _write_metadata_row(ws1, "Fixed Break (hours)", BREAK_HOURS, ATTENDANCE_BREAK_ROW)
        _write_note(
            ws1,
            "Edit Arrival Time, Departure Time, and Employee Status here via the work-hours-log-automation skill. "
            "Work Hours Calculation and Summary Report are regenerated automatically — never hand-edit those sheets.",
            len(ATTENDANCE_HEADERS),
            ATTENDANCE_NOTE_ROW,
        )

        ws2 = self._wb[SHEET_CALCULATION]
        _write_banner(ws2, f"Work Hours Log — Work Hours Calculation — {month}", len(CALCULATION_HEADERS), CALCULATION_TITLE_ROW)
        _write_note(
            ws2,
            "Derived sheet, fully regenerated from Attendance Log on every save. Confirmed Work Hours stays blank "
            "whenever Review Status is Needs Supervisor Review — never a guessed value.",
            len(CALCULATION_HEADERS),
            CALCULATION_NOTE_ROW,
        )

    def apply_data_validation_and_formatting(self) -> None:
        """Employee Status dropdown + conditional highlight for flagged rows.
        Safe to call every run — openpyxl replaces prior rules for the same ranges
        rather than stacking duplicates, since we always rebuild from a fresh load."""
        ws1 = self._wb[SHEET_ATTENDANCE]
        last_row = max(ws1.max_row, ATTENDANCE_FIRST_DATA_ROW + 500)  # headroom for future prefills
        status_col_letter = get_column_letter(_col_index(ATTENDANCE_HEADERS, "Employee Status"))
        dv = DataValidation(type="list", formula1='"Present,Leave"', allow_blank=True)
        dv.add(f"{status_col_letter}{ATTENDANCE_FIRST_DATA_ROW}:{status_col_letter}{last_row}")
        ws1.data_validations.dataValidation = []  # clear any previous instance before re-adding
        ws1.add_data_validation(dv)

        review_col_letter_1 = get_column_letter(_col_index(ATTENDANCE_HEADERS, "Review Status"))
        rule1 = CellIsRule(
            operator="equal", formula=['"Needs Supervisor Review"'],
            fill=REVIEW_HIGHLIGHT_FILL, font=REVIEW_HIGHLIGHT_FONT,
        )
        ws1.conditional_formatting.add(
            f"{review_col_letter_1}{ATTENDANCE_FIRST_DATA_ROW}:{review_col_letter_1}{last_row}", rule1
        )

        ws2 = self._wb[SHEET_CALCULATION]
        last_row_2 = max(ws2.max_row, CALCULATION_FIRST_DATA_ROW + 500)
        review_col_letter_2 = get_column_letter(_col_index(CALCULATION_HEADERS, "Review Status"))
        rule2 = CellIsRule(
            operator="equal", formula=['"Needs Supervisor Review"'],
            fill=REVIEW_HIGHLIGHT_FILL, font=REVIEW_HIGHLIGHT_FONT,
        )
        ws2.conditional_formatting.add(
            f"{review_col_letter_2}{CALCULATION_FIRST_DATA_ROW}:{review_col_letter_2}{last_row_2}", rule2
        )

    # ---- Sheet 1: Attendance Log ------------------------------------------------

    def _read_attendance_index(self) -> Dict[Tuple[str, str], int]:
        """Returns {(work_date_iso, employee_id): row_number} for existing rows."""
        ws = self._wb[SHEET_ATTENDANCE]
        date_col = _col_index(ATTENDANCE_HEADERS, "Work Date")
        id_col = _col_index(ATTENDANCE_HEADERS, "Employee ID")
        index: Dict[Tuple[str, str], int] = {}
        for row_num in range(ATTENDANCE_FIRST_DATA_ROW, ws.max_row + 1):
            d = _date_or_none(ws.cell(row=row_num, column=date_col).value)
            emp_id = ws.cell(row=row_num, column=id_col).value
            if d is None or not emp_id:
                continue
            index[(d.isoformat(), str(emp_id))] = row_num
        return index

    def prefill_month(
        self, month: str, employee_master: EmployeeMaster, holiday_calendar: HolidayCalendar
    ) -> int:
        """Ensures a blank row exists for every (active employee, Working Day date)
        in the given month, so the whole month's skeleton is visible up front, per
        the original spec. Never touches an existing row. Returns rows created."""
        year, mon = (int(part) for part in month.split("-"))
        _, days_in_month = calendar.monthrange(year, mon)
        working_dates = [
            date(year, mon, d)
            for d in range(1, days_in_month + 1)
            if holiday_calendar.day_type(date(year, mon, d)) == DayType.WORKING_DAY
        ]

        ws = self._wb[SHEET_ATTENDANCE]
        col = {name: _col_index(ATTENDANCE_HEADERS, name) for name in ATTENDANCE_HEADERS}
        index = self._read_attendance_index()
        created = 0

        for work_date in working_dates:
            for employee in employee_master.all_active():
                key = (work_date.isoformat(), employee.employee_id)
                if key in index:
                    continue
                row_num = max(ws.max_row + 1, ATTENDANCE_FIRST_DATA_ROW)
                ws.cell(row=row_num, column=col["Work Date"], value=work_date)
                ws.cell(row=row_num, column=col["Day Type"], value=DayType.WORKING_DAY.value)
                ws.cell(row=row_num, column=col["Employee ID"], value=employee.employee_id)
                ws.cell(row=row_num, column=col["Employee Name"], value=employee.employee_name)
                ws.cell(row=row_num, column=col["Department"], value=employee.department)
                ws.cell(row=row_num, column=col["Shift Start"], value=employee.shift_start)
                ws.cell(row=row_num, column=col["Shift End"], value=employee.shift_end)
                ws.cell(row=row_num, column=col["Employee Status"], value=EmployeeStatus.PRESENT.value)
                ws.cell(row=row_num, column=col["Data Source"], value=DataSource.PENDING.value)
                index[key] = row_num
                created += 1

        return created

    def upsert_attendance_rows(
        self,
        records: List[AttendanceRecord],
        employee_master: EmployeeMaster,
        holiday_calendar: HolidayCalendar,
    ) -> UpsertResult:
        ws = self._wb[SHEET_ATTENDANCE]
        result = UpsertResult()
        index = self._read_attendance_index()

        col = {name: _col_index(ATTENDANCE_HEADERS, name) for name in ATTENDANCE_HEADERS}

        for record in records:
            try:
                employee = employee_master.get(record.employee_id)
            except KeyError as exc:
                result.warnings.append(str(exc))
                continue

            key = (record.work_date.isoformat(), record.employee_id)
            day_type = holiday_calendar.day_type(record.work_date)

            if key in index:
                row_num = index[key]
                if record.arrival_time is not None:
                    ws.cell(row=row_num, column=col["Arrival Time"], value=record.arrival_time)
                if record.departure_time is not None:
                    ws.cell(row=row_num, column=col["Departure Time"], value=record.departure_time)
                ws.cell(row=row_num, column=col["Employee Status"], value=record.employee_status.value)
                ws.cell(row=row_num, column=col["Data Source"], value=record.data_source.value)
                if record.notes:
                    ws.cell(row=row_num, column=col["Notes"], value=record.notes)
                ws.cell(row=row_num, column=col["Day Type"], value=day_type.value)
                result.updated += 1
                result.touched.append(key)
            else:
                row_num = max(ws.max_row + 1, ATTENDANCE_FIRST_DATA_ROW)
                ws.cell(row=row_num, column=col["Work Date"], value=record.work_date)
                ws.cell(row=row_num, column=col["Day Type"], value=day_type.value)
                ws.cell(row=row_num, column=col["Employee ID"], value=employee.employee_id)
                ws.cell(row=row_num, column=col["Employee Name"], value=employee.employee_name)
                ws.cell(row=row_num, column=col["Department"], value=employee.department)
                ws.cell(row=row_num, column=col["Shift Start"], value=employee.shift_start)
                ws.cell(row=row_num, column=col["Shift End"], value=employee.shift_end)
                if record.arrival_time is not None:
                    ws.cell(row=row_num, column=col["Arrival Time"], value=record.arrival_time)
                if record.departure_time is not None:
                    ws.cell(row=row_num, column=col["Departure Time"], value=record.departure_time)
                ws.cell(row=row_num, column=col["Employee Status"], value=record.employee_status.value)
                ws.cell(row=row_num, column=col["Data Source"], value=record.data_source.value)
                ws.cell(row=row_num, column=col["Review Status"], value=ReviewStatus.OK.value)
                if record.notes:
                    ws.cell(row=row_num, column=col["Notes"], value=record.notes)
                index[key] = row_num
                result.added += 1
                result.touched.append(key)

        return result

    def read_all_attendance_records(self) -> List[AttendanceRecord]:
        """Reconstructs AttendanceRecords from Sheet 1's current contents, used to
        rebuild the derived sheets from source."""
        ws = self._wb[SHEET_ATTENDANCE]
        col = {name: _col_index(ATTENDANCE_HEADERS, name) for name in ATTENDANCE_HEADERS}
        records: List[AttendanceRecord] = []
        for row_num in range(ATTENDANCE_FIRST_DATA_ROW, ws.max_row + 1):
            work_date = _date_or_none(ws.cell(row=row_num, column=col["Work Date"]).value)
            employee_id = ws.cell(row=row_num, column=col["Employee ID"]).value
            if work_date is None or not employee_id:
                continue
            status_raw = ws.cell(row=row_num, column=col["Employee Status"]).value
            status = EmployeeStatus.LEAVE if status_raw == EmployeeStatus.LEAVE.value else EmployeeStatus.PRESENT
            source_raw = ws.cell(row=row_num, column=col["Data Source"]).value
            source = DataSource.CSV_IMPORT if source_raw == DataSource.CSV_IMPORT.value else DataSource.MANUAL
            records.append(
                AttendanceRecord(
                    work_date=work_date,
                    employee_id=str(employee_id),
                    arrival_time=_time_or_none(ws.cell(row=row_num, column=col["Arrival Time"]).value),
                    departure_time=_time_or_none(ws.cell(row=row_num, column=col["Departure Time"]).value),
                    employee_status=status,
                    data_source=source,
                    notes=ws.cell(row=row_num, column=col["Notes"]).value or "",
                )
            )
        return records

    def _sync_review_status_to_attendance(self, calculated_rows: List[CalculatedRow]) -> None:
        ws = self._wb[SHEET_ATTENDANCE]
        index = self._read_attendance_index()
        col = _col_index(ATTENDANCE_HEADERS, "Review Status")
        for row in calculated_rows:
            key = (row.work_date.isoformat(), row.employee_id)
            if key in index:
                ws.cell(row=index[key], column=col, value=row.review_status.value)

    # ---- Sheet 2: Work Hours Calculation (derived, always rebuilt) --------------

    def rebuild_calculation_sheet(
        self, engine: BusinessRulesEngine, employee_master: EmployeeMaster
    ) -> List[CalculatedRow]:
        ws = self._wb[SHEET_CALCULATION]
        if ws.max_row > CALCULATION_HEADER_ROW:
            ws.delete_rows(CALCULATION_FIRST_DATA_ROW, ws.max_row - CALCULATION_HEADER_ROW)

        source_records = self.read_all_attendance_records()
        calculated: List[CalculatedRow] = []
        for record in source_records:
            try:
                employee = employee_master.get(record.employee_id)
            except KeyError:
                continue  # already warned about at upsert time
            calculated.append(engine.calculate(record, employee))

        calculated.sort(key=lambda r: (r.work_date, r.employee_id))

        col = {name: _col_index(CALCULATION_HEADERS, name) for name in CALCULATION_HEADERS}
        for offset, row in enumerate(calculated, start=CALCULATION_FIRST_DATA_ROW):
            ws.cell(row=offset, column=col["Work Date"], value=row.work_date)
            ws.cell(row=offset, column=col["Employee ID"], value=row.employee_id)
            ws.cell(row=offset, column=col["Employee Name"], value=row.employee_name)
            ws.cell(row=offset, column=col["Day Type"], value=row.day_type.value)
            ws.cell(row=offset, column=col["Employee Status"], value=row.employee_status.value)
            ws.cell(row=offset, column=col["Arrival Time"], value=row.arrival_time)
            ws.cell(row=offset, column=col["Departure Time"], value=row.departure_time)
            ws.cell(row=offset, column=col["Gross Hours"], value=row.gross_hours)
            ws.cell(row=offset, column=col["Break Hours"], value=row.break_hours)
            ws.cell(row=offset, column=col["Confirmed Work Hours"], value=row.confirmed_hours)
            ws.cell(row=offset, column=col["Expected Hours"], value=row.expected_hours)
            ws.cell(row=offset, column=col["Late Arrival (Hours)"], value=row.late_arrival_hours)
            ws.cell(row=offset, column=col["Early Departure (Hours)"], value=row.early_departure_hours)
            ws.cell(row=offset, column=col["Review Status"], value=row.review_status.value)

        self._sync_review_status_to_attendance(calculated)
        return calculated

    # ---- Sheet 3: Summary Report (derived, always rebuilt) ----------------------
    # Layout intentionally unchanged from the original design — kept exactly as-is.

    def rebuild_summary_sheet(self, summary_tables: SummaryTables) -> None:
        ws = self._wb[SHEET_SUMMARY]
        for row in list(ws.iter_rows()):
            for cell in row:
                cell.value = None

        row_num = 1
        row_num = self._write_summary_table(ws, row_num, "Daily Totals", DAILY_HEADERS, [
            (r["employeeId"], r["employeeName"], r["workDate"], r["dailyWorkHours"])
            for r in summary_tables.daily
        ])
        row_num += 1
        row_num = self._write_summary_table(ws, row_num, "Weekly Totals", WEEKLY_HEADERS, [
            (r["employeeId"], r["employeeName"], r["weekStart"], r["weekEnd"], r["weeklyWorkHours"])
            for r in summary_tables.weekly
        ])
        row_num += 1
        self._write_summary_table(ws, row_num, "Monthly Totals", MONTHLY_HEADERS, [
            (r["employeeId"], r["employeeName"], r["month"], r["monthlyWorkHours"])
            for r in summary_tables.monthly
        ])

    @staticmethod
    def _write_summary_table(
        ws: Worksheet, start_row: int, title: str, headers: Tuple[str, ...], rows: List[tuple]
    ) -> int:
        title_cell = ws.cell(row=start_row, column=1, value=title)
        title_cell.font = Font(bold=True)
        header_row = start_row + 1
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
        data_row = header_row + 1
        for values in rows:
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=data_row, column=col_idx, value=value)
            data_row += 1
        return data_row  # first free row after this table

    # ---- Backup + save -----------------------------------------------------------

    def backup_if_exists(self, backup_dir: Path) -> Optional[Path]:
        if not self.workbook_path.exists():
            return None
        backup_dir.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%dT%H%M%S")
        backup_path = backup_dir / f"{self.workbook_path.stem}__backup-{timestamp}.xlsx"
        shutil.copy2(self.workbook_path, backup_path)
        return backup_path

    def raise_if_modified_since_load(self) -> None:
        """Optimistic-concurrency guard: call this right before save(). If another
        process (e.g. a different PC on a shared drive) wrote to this workbook
        after this instance loaded it, raise rather than silently overwrite."""
        exists_now = self.workbook_path.exists()
        if self._loaded_hash is None:
            if exists_now:
                raise ConcurrentModificationError(
                    f"{self.workbook_path} was created by someone else since this "
                    "session started. Re-run to load the current version before saving."
                )
            return
        if not exists_now:
            raise ConcurrentModificationError(
                f"{self.workbook_path} was deleted by someone else since this "
                "session started. Re-run to confirm the current state before saving."
            )
        if self._file_hash(self.workbook_path) != self._loaded_hash:
            raise ConcurrentModificationError(
                f"{self.workbook_path} was changed by someone else since this "
                "session started. Re-run to load their changes before saving yours."
            )

    def save(self) -> None:
        self.workbook_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(str(self.workbook_path))
